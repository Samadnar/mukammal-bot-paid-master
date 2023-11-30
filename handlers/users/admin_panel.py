from aiogram import types
from aiogram.types import ContentType
from loader import dp, db_book
from aiogram.dispatcher import FSMContext
from states.add_book import Add_book
from filters.admin_filtr import Adminfiltr
from aiogram.dispatcher.filters import Command
from data.config import ADMINS
from openpyxl import Workbook

@dp.message_handler(Command("add_books", prefixes="/"), user_id=ADMINS)
async def insert_book_name(message: types.Message):
    await message.answer("Kitob nomi")
    await Add_book.book_name.set()

@dp.message_handler(state=Add_book.book_name)
async def insert_number_of_books(message: types.Message, state: FSMContext):
    await state.update_data(
        {"book_name": message.text}
    )
    await message.answer("Kitob soni")
    await Add_book.number_of_books.set()
    
@dp.message_handler(state=Add_book.number_of_books)
async def insert_price_the_book(message: types.Message, state: FSMContext):
    await state.update_data(
        {"number_of_books": int(message.text)}
    )
    await message.answer("Kitob narxi")
    await Add_book.price_the_book.set()
    
@dp.message_handler(state=Add_book.price_the_book)
async def insert_book_image(message: types.Message, state: FSMContext):
    await state.update_data(
        {"price_the_book": int(message.text)}
    )
    await message.answer("Kitob rasmi")
    await Add_book.image.set()
    
@dp.message_handler(state=Add_book.image, content_types=ContentType.PHOTO)
async def insert_writer_name(message: types.Message, state: FSMContext):
    await state.update_data(
        {"image": message.photo[-1].file_id}
    )
    await message.answer("Kitob muallifi")
    await Add_book.writer_name.set()
    
@dp.message_handler(state=Add_book.writer_name)
async def insert_type_book(message: types.Message, state: FSMContext):
    await state.update_data(
        {"writer_name": message.text}
    )
    await message.answer("Kitob janri\nBadiiy adabiyotlar -> 1\nBiznes va shaxsiy rivojlanish -> 2\nBolalar adabiyoti -> 3\nDiniy va ma'rifiy adabiyotlar -> 4\nFoydali adabiyotlar -> 5\nIlmiy adabiyotlar -> 6")
    "Badiiy adabiyotlar -> 1"
    await Add_book.types.set()
    
@dp.message_handler(state=Add_book.types)
async def get_type_book(message: types.Message, state: FSMContext):
    
    await state.update_data(
        {"book_type": message.text}
    )

    await message.answer("Kitob qo'shildi")
    data = await state.get_data()
    book_name = data.get("book_name")
    number_of_books = data.get("number_of_books")
    price_the_book = data.get("price_the_book")
    image = data.get("image")
    writer_name = data.get("writer_name")
    type = data.get("book_type")
    await db_book.add_books(book_name, number_of_books, price_the_book,  image, writer_name, type)
    await state.finish()


@dp.message_handler(Command("view_users", prefixes="/"), user_id=ADMINS)
async def insert_book_name(message: types.Message):
    all_users = await db_book.select_all_users()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Telegram id"
    ws["B1"] = "Familiya Ism"
    ws["C1"] = "Telefon raqam"
    for i in range(0, len(all_users)):
        ws[f"A{i+2}"] = all_users[i][1]
        ws[f"B{i+2}"] = all_users[i][2]
        ws[f"C{i+2}"] = all_users[i][3]
    exl_file = wb.save("all_users.xlsx")
    await message.answer_document(open("all_users.xlsx", 'rb'))